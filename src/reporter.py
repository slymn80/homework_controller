# src/reporter.py
from __future__ import annotations
from pathlib import Path
from typing import List, Dict, Any
import json
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

HEADERS = [
    "first_name",
    "last_name",
    "class",
    "student",
    "file_name",
    "file_id",
    "word_count",
    "total",
    "content",
    "structure",
    "language",
    "originality",
    "feedback",
]

def _cell(value: Any) -> Any:
    """Excel hücresi için güvenli değer: dict/list -> JSON string, None -> '', diğerleri aynen."""
    if value is None:
        return ""
    if isinstance(value, (dict, list)):
        # unicode korunsun diye ensure_ascii=False
        return json.dumps(value, ensure_ascii=False)
    return value

def create_report_excel(output_path: str, rows: List[Dict[str, Any]]) -> str:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"

    # Başlıklar
    ws.append(HEADERS)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    for col, _h in enumerate(HEADERS, 1):
        c = ws.cell(row=1, column=col)
        c.font = header_font
        c.fill = header_fill
        ws.column_dimensions[c.column_letter].width = 20
    ws.column_dimensions["M"].width = 80  # feedback daha geniş

    # Satırlar
    for r in rows:
        bd = r.get("breakdown") or {}
        ws.append([
            _cell(r.get("first_name")),
            _cell(r.get("last_name")),
            _cell(r.get("class")),
            _cell(r.get("student")),
            _cell(r.get("file_name")),
            _cell(r.get("file_id")),
            _cell(r.get("word_count")),
            _cell(r.get("total")),
            _cell(bd.get("content", r.get("content"))),
            _cell(bd.get("structure", r.get("structure"))),
            _cell(bd.get("language", r.get("language"))),
            _cell(bd.get("originality", r.get("originality"))),
            _cell(r.get("feedback")),
        ])

    # Hücre biçimleri
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = Alignment(wrap_text=True, vertical="top")

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
